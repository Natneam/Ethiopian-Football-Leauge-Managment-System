from tkinter import*
import openpyxl
from tkinter import ttk
from tkinter import messagebox as mBox
class Week:
   def __init__(self,width,width2,win,screen,look):
      self.width=width
      self.screen=screen
      self.look=look
      self.win=win
      self.width2=width2
   def week1(self,width,width2,win,screen,look):
      self.win = Tk()
      self.win.resizable(0,0)
      self.win.title("Ethiopian Football League  Management System")
      self.win.geometry("900x500")
      self.win.resizable(0,0)
      
      
      frame = Frame(self.win)
      Label(frame,text="MATCH FIXTURES").grid(row=0,columnspan=6)
      Label(frame,text="WEEK ONE").grid(row=1,columnspan=6,pady=25)
       
      Label(frame,text="Sidama Bunna").grid(row=2,column=0,sticky=E)
      self.entry1=Entry(frame,width=self.width)
      self.entry1.grid(row=2,column=1)
      Label(frame,text="Vs").grid(row=2,column=2)
      self.entry2=Entry(frame,width=self.width)
      self.entry2.grid(row=2,column=3)
      Label(frame,text="Fasil Ketema").grid(row=2,column=4,sticky=W)
      Label(frame,text="").grid(row=2,column=0)


      Label(frame,text="Mekelle 70 enderta F.C").grid(row=3,column=0,sticky=E)
     
      self.entry3=Entry(frame,width=self.width)
      self.entry3.grid(row=3,column=1)
      Label(frame,text="Vs").grid(row=3,column=2)
     
      self.entry4=Entry(frame,width=self.width)
      self.entry4.grid(row=3,column=3)
      Label(frame,text="Dedebit").grid(row=3,column=4,pady=8,sticky=W)
       
       
      Label(frame,text="Hawassa").grid(row=4,column=0,sticky=E)

      self.entry5=Entry(frame,width=self.width)
      self.entry5.grid(row=4,column=1)
      Label(frame,text="Vs").grid(row=4,column=2)
 
      self.entry6=Entry(frame,width=self.width)
      self.entry6.grid(row=4,column=3)
      Label(frame,text="wolewalo adigrat").grid(row=4,column=4,pady=8,sticky=W)


      Label(frame,text="Adama city").grid(row=5,column=0,sticky=E)

      self.entry7=Entry(frame,width=self.width)
      self.entry7.grid(row=5,column=1)
      Label(frame,text="Vs").grid(row=5,column=2)
      
      self.entry8=Entry(frame,width=self.width)
      self.entry8.grid(row=5,column=3)
      Label(frame,text="Jimma Aba Jifar ").grid(row=5,column=4,pady=8,sticky=W)

      Label(frame,text="Ethiopia Bunna").grid(row=6,column=0,sticky=E)
   
      self.entry9=Entry(frame,width=self.width)
      self.entry9.grid(row=6,column=1)
      Label(frame,text="Vs").grid(row=6,column=2)
  
      self.entry10=Entry(frame,width=self.width)
      self.entry10.grid(row=6,column=3)
      Label(frame,text="Dire Dawa ").grid(row=6,column=4,pady=8,sticky=W)


      Label(frame,text="Shire Endeslassie").grid(row=7,column=0,sticky=E)
      self.entry11=Entry(frame,width=self.width)
      self.entry11.grid(row=7,column=1)
      Label(frame,text="Vs").grid(row=7,column=2)
  
      self.entry12=Entry(frame,width=self.width)
      self.entry12.grid(row=7,column=3)
      Label(frame,text="Welyata Dicha").grid(row=7,column=4,pady=8,sticky=W)
      Label(frame,text="Debub Police").grid(row=8,column=0,sticky=E)
      self.var13=IntVar
      self.entry13=Entry(frame,width=self.width)
      self.entry13.grid(row=8,column=1)
      Label(frame,text="Vs").grid(row=8,column=2)
      
      self.entry14=Entry(frame,width=self.width)
      self.entry14.grid(row=8,column=3)
      Label(frame,text="Defence Force").grid(row=8,column=4,pady=8,sticky=W)

      Label(frame,text="st.George").grid(row=9,column=0,sticky=E)
  
      self.entry15=Entry(frame,width=self.width)
      self.entry15.grid(row=9,column=1)
      Label(frame,text="Vs").grid(row=9,column=2)
      
      self.entry16=Entry(frame,width=self.width)
      self.entry16.grid(row=9,column=3)
      Label(frame,text="Bahir Dar Kenema").grid(row=9,column=4,pady=8,sticky=W)
      Button(frame,text="Next",command=self.destroy).grid(row=15,column=6,sticky=E)
      Button(frame,text="Submit",command=self.get,width=self.width2).grid(row=15,column=4,sticky=E)
      frame.grid(padx=270,pady=10)
      self.win.mainloop()
   def week2(self,width,width2,win,screen,look):
    self.screen = Tk()
    self.screen.resizable(0,0)
    self.screen.title("Ethiopian Football League  Management System")
    self.screen.geometry("900x500")
    self.screen.resizable(0,0)
    frame2=Frame(self.screen)
    Label(frame2,text="MATCH FIXTURES").grid(row=0,columnspan=6)
    Label(frame2,text="WEEK TWO").grid(row=1,columnspan=6,pady=25)
    
    Label(frame2,text="wolewalo adigrat").grid(row=2,column=0,sticky=E)
    self.input1=Entry(frame2,width=self.width)
    self.input1.grid(row=2,column=1)
    Label(frame2,text="Vs").grid(row=2,column=2)
    self.input2=Entry(frame2,width=self.width)
    self.input2.grid(row=2,column=3)
    Label(frame2,text="Adama city").grid(row=2,column=4,sticky=W)
    Label(frame2,text="").grid(row=2,column=0)


    Label(frame2,text="Dedebit").grid(row=3,column=0,sticky=E)
    self.input3=Entry(frame2,width=self.width)
    self.input3.grid(row=3,column=1)
    Label(frame2,text="Vs").grid(row=3,column=2)
    self.input4=Entry(frame2,width=self.width)
    self.input4.grid(row=3,column=3)
    Label(frame2,text="Ethiopia Bunna").grid(row=3,column=4,pady=8,sticky=W)
    
    
    Label(frame2,text="Fasil Ketema").grid(row=4,column=0,sticky=E)
    self.input5=Entry(frame2,width=self.width)
    self.input5.grid(row=4,column=1)
    Label(frame2,text="Vs").grid(row=4,column=2)
    self.input6=Entry(frame2,width=self.width)
    self.input6.grid(row=4,column=3)
    Label(frame2,text="Hawassa ").grid(row=4,column=4,pady=8,sticky=W)


    Label(frame2,text="Dire Dawa").grid(row=5,column=0,sticky=E)
    self.input7=Entry(frame2,width=self.width)
    self.input7.grid(row=5,column=1)
    Label(frame2,text="Vs").grid(row=5,column=2)
    self.input8=Entry(frame2,width=5)
    self.input8.grid(row=5,column=3)
    Label(frame2,text="Debub Police").grid(row=5,column=4,pady=8,sticky=W)

    Label(frame2,text="Adama city").grid(row=6,column=0,sticky=E)
    self.input9=Entry(frame2,width=self.width)
    self.input9.grid(row=6,column=1)
    Label(frame2,text="Vs").grid(row=6,column=2)
    self.input10=Entry(frame2,width=self.width)
    self.input10.grid(row=6,column=3)
    Label(frame2,text="st.George").grid(row=6,column=4,pady=8,sticky=W)
    Label(frame2,text="Welyata Dicha").grid(row=7,column=0,sticky=E)
    self.input11=Entry(frame2,width=5)
    self.input11.grid(row=7,column=1)
    Label(frame2,text="Vs").grid(row=7,column=2)
    self.input12=Entry(frame2,width=self.width)
    self.input12.grid(row=7,column=3)
    Label(frame2,text="Sidama Bunna").grid(row=7,column=4,pady=8,sticky=W)
    
    
    Label(frame2,text="Bahir Dar Kenema").grid(row=8,column=0,sticky=E)
    self.input13=Entry(frame2,width=self.width)
    self.input13.grid(row=8,column=1)
    Label(frame2,text="Vs").grid(row=8,column=2)
    self.input14=Entry(frame2,width=self.width)
    self.input14.grid(row=8,column=3)
    Label(frame2,text="Shire Endeslassie").grid(row=8,column=4,pady=8,sticky=W)


    Label(frame2,text="Jimma Aba Jifar ").grid(row=9,column=0,sticky=E)
    self.input15=Entry(frame2,width=self.width)
    self.input15.grid(row=9,column=1)
    
    Label(frame2,text="Vs").grid(row=9,column=2)
    self.input16=Entry(frame2,width=self.width)
    self.input16.grid(row=9,column=3)
    Label(frame2,text="Defence Force ").grid(row=9,column=4,pady=8,sticky=W)
    Button(frame2,text="Previous",command=self.destroy4).grid(row=15,column=5,sticky=E)
    Button(frame2,text="Next",command=self.destroy2).grid(row=15,column=6,sticky=E)
    Button(frame2,text="Submit",command=self.get1,width=self.width2).grid(row=15,column=4,sticky=E)
    frame2.grid(padx=270,pady=10)
    self.screen.mainloop()
   def week3(self,width,width2,win,screen,look):

       self.look = Tk()
       self.look.resizable(0,0)
       self.look.title("Ethiopian Football League  Management System")
       self.look.geometry("900x500")
       self.look.resizable(0,0)
       frame3=Frame(self.look)
       Label(frame3,text="MATCH FIXTURES").grid(row=0,columnspan=6)
       Label(frame3,text="WEEK THREE").grid(row=1,columnspan=6,pady=25)
       
       Label(frame3,text="Defence Force ").grid(row=2,column=0,sticky=E)
       self.user1=Entry(frame3,width=self.width)
       self.user1.grid(row=2,column=1)
       Label(frame3,text="Vs").grid(row=2,column=2)
       self.user2=Entry(frame3,width=self.width)
       self.user2.grid(row=2,column=3)
       Label(frame3,text="Dire Dawa ").grid(row=2,column=4,sticky=W)
       Label(frame3,text="").grid(row=2,column=0)


       Label(frame3,text="Debub Police").grid(row=3,column=0,sticky=E)
       self.user3=Entry(frame3,width=self.width)
       self.user3.grid(row=3,column=1)
       Label(frame3,text="Vs").grid(row=3,column=2)
       self.user4=Entry(frame3,width=self.width)
       self.user4.grid(row=3,column=3)
       Label(frame3,text="Dedebit").grid(row=3,column=4,pady=8,sticky=W)
       
       
       Label(frame3,text="Sidama Bunna").grid(row=4,column=0,sticky=E)
       self.user5=Entry(frame3,width=self.width)
       self.user5.grid(row=4,column=1)
       Label(frame3,text="Vs").grid(row=4,column=2)

       self.user6=Entry(frame3,width=self.width)
       self.user6.grid(row=4,column=3)
       Label(frame3,text="Bahir Dar Kenema").grid(row=4,column=4,pady=8,sticky=W)


       Label(frame3,text="Adama city").grid(row=5,column=0,sticky=E)
       self.user7= Entry(frame3,width=self.width)
       self.user7.grid(row=5,column=1)
       Label(frame3,text="Vs").grid(row=5,column=2)
       self.user8=Entry(frame3,width=self.width)
       self.user8.grid(row=5,column=3)
       Label(frame3,text="Fasil Ketema").grid(row=5,column=4,pady=8,sticky=W)

       Label(frame3,text="Hawassa").grid(row=6,column=0,sticky=E)
       self.user9=Entry(frame3,width=5)
       self.user9.grid(row=6,column=1)
       Label(frame3,text="Vs").grid(row=6,column=2)
       self.user10=Entry(frame3,width=self.width)
       self.user10.grid(row=6,column=3)
       Label(frame3,text="Welyata Dicha ").grid(row=6,column=4,pady=8,sticky=W)


       Label(frame3,text="Shire Endeslassie").grid(row=7,column=0,sticky=E)
       self.user11=Entry(frame3,width=self.width)
       self.user11.grid(row=7,column=1)
       Label(frame3,text="Vs").grid(row=7,column=2)
       self.user12=Entry(frame3,width=self.width)
       self.user12.grid(row=7,column=3)
       Label(frame3,text="Adama city").grid(row=7,column=4,pady=8,sticky=W)
   
       Label(frame3,text="st.George",).grid(row=8,column=0,sticky=E)
       self.user13=Entry(frame3,width=self.width)
       self.user13.grid(row=8,column=1)
       Label(frame3,text="Vs").grid(row=8,column=2)
       self.user14=Entry(frame3,width=self.width)
       self.user14.grid(row=8,column=3)
       Label(frame3,text="Jimma Aba Jifar").grid(row=8,column=4,pady=8,sticky=W)


       Label(frame3,text="Ethiopia Bunna").grid(row=9,column=0,sticky=E)
       self.user15=Entry(frame3,width=self.width)
       self.user15.grid(row=9,column=1)
       Label(frame3,text="Vs").grid(row=9,column=2)
       self.user16=Entry(frame3,width=self.width)
       self.user16.grid(row=9,column=3)
       Label(frame3,text="wolewalo adigrat").grid(row=9,column=4,pady=8,sticky=W)
       Button(frame3,text="Previous",command=self.destroy3,width=self.width2).grid(row=15,column=6,sticky=E)
       Button(frame3,text="Submit",command=self.get2,width=self.width2).grid(row=15,column=4,sticky=E)
       frame3.grid(padx=270,pady=10)
       self.look.mainloop()   

   def playerError(self):
      mBox.showinfo('goal scorer info box','you must fill the entries correctly') 
           
      
   
      
      
   def write1(self):
      wb = openpyxl.load_workbook('league.xlsx')
      sheet = wb["Fixture"]
      sheet.cell(row = 5, column = 4).value = self.var1
      sheet.cell(row = 5, column = 5).value = self.var2
      sheet.cell(row = 6, column = 4).value = self.var3
      sheet.cell(row = 6, column = 5).value = self.var4
      sheet.cell(row = 7, column = 4).value = self.var5
      sheet.cell(row = 7, column = 5).value = self.var6
      sheet.cell(row = 8, column = 4).value = self.var7
      sheet.cell(row = 8, column = 5).value = self.var8
      sheet.cell(row = 9, column = 4).value = self.var9
      sheet.cell(row = 9, column = 5).value = self.var10
      sheet.cell(row = 10, column = 4).value = self.var11
      sheet.cell(row = 10, column = 5).value = self.var12
      sheet.cell(row = 11, column = 4).value = self.var13
      sheet.cell(row = 11, column = 5).value = self.var14
      sheet.cell(row = 12, column = 4).value = self.var15
      sheet.cell(row = 12, column = 5).value = self.var16
      wb.save('league.xlsx')
   def write2(self):
      wb = openpyxl.load_workbook('league.xlsx')
      sheet = wb["Fixture"]
      sheet.cell(row = 13, column = 4).value = self.x1
      sheet.cell(row = 13, column = 5).value = self.x2
      sheet.cell(row = 14, column = 4).value = self.x3
      sheet.cell(row = 14, column = 5).value = self.x4
      sheet.cell(row = 15, column = 4).value = self.x5
      sheet.cell(row = 15, column = 5).value = self.x6
      sheet.cell(row = 16, column = 4).value = self.x7
      sheet.cell(row = 16, column = 5).value = self.x8
      sheet.cell(row = 17, column = 4).value = self.x9
      sheet.cell(row = 17, column = 5).value = self.x10
      sheet.cell(row = 18, column = 4).value = self.x11
      sheet.cell(row = 18, column = 5).value = self.x12
      sheet.cell(row = 19, column = 4).value = self.x13
      sheet.cell(row = 19, column = 5).value = self.x14
      sheet.cell(row = 20, column = 4).value = self.x15
      sheet.cell(row = 20, column = 5).value = self.x16
      wb.save('league.xlsx')
   def write3(self):
      wb = openpyxl.load_workbook('league.xlsx')
      sheet = wb["Fixture"]
      sheet.cell(row = 21, column = 4).value = self.y1
      sheet.cell(row = 21, column = 5).value = self.y2
      sheet.cell(row = 22, column = 4).value = self.y3
      sheet.cell(row = 22, column = 5).value = self.y4
      sheet.cell(row = 23, column = 4).value = self.y5
      sheet.cell(row = 23, column = 5).value = self.y6
      sheet.cell(row = 24, column = 4).value = self.y7
      sheet.cell(row = 24, column = 5).value = self.y8
      sheet.cell(row = 25, column = 4).value = self.y9
      sheet.cell(row = 25, column = 5).value = self.y10
      sheet.cell(row = 26, column = 4).value = self.y11
      sheet.cell(row = 26, column = 5).value = self.y12
      sheet.cell(row = 27, column = 4).value = self.y13
      sheet.cell(row = 27, column = 5).value = self.y14
      sheet.cell(row = 28, column = 4).value = self.y15
      sheet.cell(row = 28, column = 5).value = self.y16
      wb.save('league.xlsx')

   def error(self):
      self.answer=mBox.askyesno("ENTRY MESSAGE","You must Fill  all necessary entries\n There will be problem if you continue.\nDo you want to continue in any way?")
      
   def get(self):
      self.var1=self.entry1.get()
      self.var2=self.entry2.get()
      self.var3=self.entry3.get()
      self.var4=self.entry4.get()
      self.var5=self.entry5.get()
      self.var6=self.entry6.get()
      self.var7=self.entry7.get()
      self.var8=self.entry8.get()
      self.var9=self.entry9.get()
      self.var10=self.entry10.get()
      self.var11=self.entry11.get()
      self.var12=self.entry12.get()
      self.var13=self.entry13.get()
      self.var14=self.entry14.get()
      self.var15=self.entry15.get()
      self.var16=self.entry16.get()
      if (self.var1!="" and self.var2=="") or (self.var2!="" and self.var1==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass  
      elif (self.var3!="" and self.var4=="") or (self.var4!="" and self.var3==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      elif (self.var5!="" and self.var6=="") or (self.var6!="" and self.var5==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      elif (self.var7!="" and self.var8=="") or (self.var8!="" and self.var7==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      elif (self.var9!="" and self.var10=="") or (self.var10!="" and self.var9==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      elif (self.var11!="" and self.var12=="") or (self.var12!="" and self.var11==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      elif (self.var13!="" and self.var14=="") or (self.var14!="" and self.var13==""):
         self.error()
         if self.answer==True:
            self.self.write1()
         else:
            pass
      elif (self.var15!="" and self.var16=="") or (self.var16!="" and self.var15==""):
         self.error()
         if self.answer==True:
            self.write1()
         else:
            pass
      else:
         self.write1()
   def get1(self):
      self. x1=self.input1.get()
      self.x2=self.input2.get()
      self.x3=self.input3.get()
      self.x4=self.input4.get()
      self.x5=self.input5.get()
      self.x6=self.input6.get()
      self.x7=self.input7.get()
      self.x8=self.input8.get()
      self.x9=self.input9.get()
      self.x10=self.input10.get()
      self.x11=self.input11.get()
      self.x12=self.input12.get()
      self.x13=self.input13.get()
      self.x14=self.input14.get()
      self.x15=self.input15.get()
      self.x16=self.input16.get()
      
      if (self.x1!="" and self.x2=="") or (self.x2!="" and self.x1==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass
      elif (self.x3!="" and self.x4=="") or (self.x4!="" and self.x3==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x5!="" and self.x6=="") or (self.x6!="" and self.x5==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x7!="" and self.x8=="") or (self.x8!="" and self.x7==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x9!="" and self.x10=="") or (self.x10!="" and self.x9==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x11!="" and self.x12=="") or (self.x12!="" and self.x11==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x13!="" and self.x14=="") or (self.x14!="" and self.x13==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      elif (self.x15!="" and self.x16=="") or (self.x16!="" and self.x15==""):
         self.error()
         if self.answer==True:
            self.write2()
         else:
            pass

      else:
         self.write2()
      if self.answer==True:
         self.write2()
      else:
         pass
   def get2(self):
      self.y1=self.user1.get()
      self.y2=self.user2.get()
      self.y3=self.user3.get()
      self.y4=self.user4.get()
      self.y5=self.user5.get()
      self.y6=self.user6.get()
      self.y7=self.user7.get()
      self.y8=self.user8.get()
      self.y9=self.user9.get()
      self.y10=self.user10.get()
      self.y11=self.user11.get()
      self.y12=self.user12.get()
      self.y13=self.user13.get()
      self.y14=self.user14.get()
      self.y15=self.user15.get()
      self.y16=self.user16.get()
      
      if (self.y1!="" and self.y2=="") or (self.y2!="" and self.y1==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y3!="" and self.y4=="") or (self.y4!="" and self.y3==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y5!="" and self.y6=="") or (self.y6!="" and self.y5==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y7!="" and self.y8=="") or (self.y8!="" and self.y7==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y9!="" and self.y10=="") or (self.y10!="" and self.y9==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y11!="" and self.y12=="") or (self.y12!="" and self.y11==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y13!="" and self.y14=="") or (self.y14!="" and self.y13==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      elif (self.y15!="" and self.y16=="") or (self.y16!="" and self.y15==""):
         self.error()
         if self.answer==True:
            self.write3()
         else:
            pass
      else:
         self.write3()
      if self.answer==True:
         self.write3()
      else:
         pass
   def destroy(self):
       self.win.destroy()
       self.week2(self.width,self.width2,self.win,self.screen,self.look)
   def destroy2(self):
       self.screen.destroy()
       self.week3(self.width,self.width2,self.win,self.screen,self.look)
   def destroy3(self):
       self.look.destroy()
       self.week2(self.width,self.width2,self.win,self.screen,self.look)
   def destroy4(self):
       self.screen.destroy()
       self.week1(self.width,self.width2,self.win,self.screen,self.look)
         

def week1():
   obj=Week(4,12,"win","screen","look")
   return obj.week1(4,12,"win","screen","look")
   




   
