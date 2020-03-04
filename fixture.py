from tkinter import*
import openpyxl
def destroy():
    win.destroy()
    fweek2()
def destroy2():
    screen.destroy()
    fweek3()
def destroy3():
    look.destroy()
    fweek2()
def destroy4():
    screen.destroy()
    fweek1()

def fweek3():
    global look
    look=Tk()
    frame2=Frame(look)
    look.title("Ethiopian Football League  Management System")
    wb = openpyxl.load_workbook('league.xlsx')
    sheet = wb["Fixture"]
    Label(frame2,text="MATCH RESULTS").grid(row=2,columnspan=8,pady=10)
    Label(frame2,text="WEEK THREE").grid(row=3,columnspan=8,pady=10)
    Label(frame2,text=str(sheet.cell(row=4,column=2).value)).grid(row=4,column=2,sticky=E)
    Label(frame2,text=str(sheet.cell(row=4,column=3).value)).grid(row=4,column=6,sticky=W)
    for i in range(21,29):
            for k in range(2,4):
                if k==2:
                    Label(frame2,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=2,sticky=E)
                    Label(frame2,text=str(sheet.cell(row=i,column=4).value)).grid(row=i,column=3)
                    Label(frame2,text="-").grid(row=i,column=4)
                else:
                    Label(frame2,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=6,sticky=W)
                    Label(frame2,text=str(sheet.cell(row=i,column=5).value)).grid(row=i,column=5)
    Button(frame2,text="Previous",command=destroy3).grid(row=29,column=  10)    
    frame2.grid(padx=300,pady=100)
    look.mainloop()
    wb.save('league.xlsx')
def fweek2():
    global screen
    screen=Tk()
    screen.title("Ethiopian Football League  Management System")
    frame1=Frame(screen)
    wb = openpyxl.load_workbook('league.xlsx')
    sheet = wb["Fixture"]
    Label(frame1,text="MATCH RESULTS").grid(row=2,columnspan=8,pady=10)
    Label(frame1,text="WEEK TWO").grid(row=3,columnspan=8,pady=10)
    Label(frame1,text=str(sheet.cell(row=4,column=2).value)).grid(row=4,column=2,sticky=E)
    Label(frame1,text=str(sheet.cell(row=4,column=3).value)).grid(row=4,column=6,sticky=W)
    for i in range(13,21):
            for k in range(2,4):
                if k==2:
                    Label(frame1,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=2,sticky=E)
                    Label(frame1,text=str(sheet.cell(row=i,column=4).value)).grid(row=i,column=3)
                    Label(frame1,text="-").grid(row=i,column=4)
                else:
                    Label(frame1,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=6,sticky=W)
                    Label(frame1,text=str(sheet.cell(row=i,column=5).value)).grid(row=i,column=5)
    Button(frame1,text="Previous",command=destroy4).grid(row=21,column=10)
    Button(frame1,text="Next",command=destroy2).grid(row=21,column=11)     
    frame1.grid(padx=300,pady=100)
    screen.mainloop()
    wb.save('league.xlsx')
def fweek1():
    global win
    win=Tk()
    win.title("Ethiopian Football League  Management System")
    frame=Frame(win)
    wb = openpyxl.load_workbook('league.xlsx')
    sheet = wb["Fixture"]
    Label(frame,text="MATCH RESULTS").grid(row=2,columnspan=8,pady=10)
    Label(frame,text="WEEK ONE").grid(row=3,columnspan=8,pady=10)
    Label(frame,text=str(sheet.cell(row=4,column=2).value)).grid(row=4,column=2,sticky=E)
    Label(frame,text=str(sheet.cell(row=4,column=3).value)).grid(row=4,column=6,sticky=W)
    for i in range(5,13):
        for k in range(2,4):
            if k==2:
                Label(frame,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=2,sticky=E)
                Label(frame,text=str(sheet.cell(row=i,column=4).value)).grid(row=i,column=3)
                Label(frame,text="-").grid(row=i,column=4)
            else:
                Label(frame,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=6)
                Label(frame,text=str(sheet.cell(row=i,column=5).value)).grid(row=i,column=5)
    Button(frame,text="Next",command=destroy).grid(row=13,columnspan=8,sticky=E)     
    frame.grid(padx=300,pady=100)
    win.mainloop()
    wb.save('league.xlsx')

