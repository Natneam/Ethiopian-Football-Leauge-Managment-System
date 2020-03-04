import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox as mBox
wb = openpyxl.load_workbook("league.xlsx")


def add():
    
    club = entry_club.get()
    club = club.lower()
    name = entry_player.get()
    name = name.lower()
    while True:
        try:
             number = int(entry_jersey.get())
             break
        except:   
            mBox.showinfo('error box', 'enter an integer')
            return insert
    while True:
        try:
             age = int(entry_age.get())
             break
        except:   
            mBox.showinfo('error box', 'enter an integer greater than 1')
            return insert
    role = entry_role.get()

    i = 2 
    j = 1
    
    sheet = wb[club]
    while sheet['A' + str(i)].value != None :
        i += 1
        
    sheet.cell(row = i, column = j).value = name
    sheet.cell(row = i, column = j + 1).value = number
    sheet.cell(row = i, column = j + 2).value = age
    sheet.cell(row = i, column = j + 3).value = role
    mBox.showinfo('player adding msg box', 'Player Added successfuly')
    return wb.save('league.xlsx')
 
def delete():
    
    club = entry1_club.get()
    club = club.lower()
    name = entry1_player.get()
    name = name.lower()        
    i = 2
    j = 1
    sheet = wb[club]
    while sheet['A' + str(i)].value != name:
        i += 1
    answer = mBox.askyesno('Dual choice Box', 'Are you sure you really wish to delete this player ')
    if answer == True: 
        sheet.cell(row = i, column = j).value = None
        sheet.cell(row = i, column = j + 1).value = None
        sheet.cell(row = i, column = j + 2).value = None
        sheet.cell(row = i, column = j + 3).value = None
        mBox.showinfo('player deleteing msg box', 'Player Delete successfuly')
        return wb.save('league.xlsx')
    else:
        pass

def transfer():
   
    T_from = entry_from.get()
    T_from = T_from.lower()
    T_to = entry_to.get()
    T_to = T_to.lower()
    name = entry_name.get()
    name = name.lower()
    while True:
        try:
            cash = entry_cash.get()
            cash = cash.lower()
            break
        except:   
            mBox.showinfo('error box', 'enter an integer greater than 1')
            return insert
    i = 2 
    j = 1
    sheet = wb["transfer"]
    while sheet['A' + str(i)].value != None :
        i += 1  
    sheet.cell(row = i, column = j).value = name 
    sheet.cell(row = i, column = j + 1).value = T_from
    sheet.cell(row = i, column = j + 2).value =T_to 
    sheet.cell(row = i, column = j + 3).value = cash
    mBox.showinfo('player transfer msg box', 'Player Transfer successfuly')
    return wb.save('league.xlsx')

##delete player
    i = 2
    j = 1
    sheet = wb[club]
    while sheet['A' + str(i)].value != name:
        i += 1
    answer = mBox.askyesno('Dual choice Box', 'Are you sure you really wish to delete this player ')
    if answer == True: 
        sheet.cell(row = i, column = j).value = None
        sheet.cell(row = i, column = j + 1).value = None
        sheet.cell(row = i, column = j + 2).value = None
        sheet.cell(row = i, column = j + 3).value = None
        mBox.showinfo('player deleteing msg box', 'Player Delete successfuly')
        return wb.save('league.xlsx')
    else:
        pass


    

    
def insert():
    window = Tk()
    canvas3 = Canvas(window, bg = 'white')
    title3 = Label(canvas3, text = 'Transfer Form')
    title3.grid(row =6 , column =1, pady = 10)
    global entry_from
    Tfrom = Label(canvas3, text = 'Transfer From')
    Tfrom.grid(row = 7, column = 0, pady = 10)
    
    entry_from = ttk.Combobox(canvas3, width = 12, state = 'readonly')
    entry_from['values'] = ('adama City' , 'bahir dar kenema','debub police','dedebit','defence force','dire dawa','ethiopia bunna','fasil ketema',
             'hawassa','jimma aba jifar','mekelle 70 enderta f.c','shire endeslassie','sidama bunna','st.george','welyata dicha',
             'welewalo adigrat')
    
    entry_from.grid(row = 7, column = 1)
    global entry_to
    Tto = Label(canvas3, text = 'Transfer To')
    Tto.grid(row = 8, column = 0, pady = 10)
    To = StringVar
    entry_to = ttk.Combobox(canvas3, width = 12, textvariable = To)
    entry_to['values'] = ('adama city' , 'bahir dar kenema','debub police','dedebit','defence force','dire dawa','ethiopia bunna','fasil ketema',
             'hawassa','jimma aba jifar','mekelle 70 enderta f.c','shire endeslassie','sidama bunna','st.george','welyata dicha',
             'welewalo adigrat')
    entry_to.grid(row = 8, column = 1)
    global entry_name
    player_name = Label(canvas3, text = 'Name of player')
    player_name.grid(row = 9, column = 0, pady = 10)
    entry_name = Entry(canvas3)
    entry_name.grid(row = 9, column = 1)
    global entry_cash
    cash = Label(canvas3, text = 'Transfer Cash')
    cash.grid(row = 11, column = 0, pady = 10)
    entry_cash = Entry(canvas3)
    entry_cash.grid(row = 11, column = 1, padx = 15)
    submit_btn3 = Button(canvas3, text = 'TRANSFER', command = transfer)
    submit_btn3.grid(row = 11, column = 2, pady = 10)


    canvas2 = Canvas(window, bg = 'white')
    title1 = Label(canvas2, text = 'To ADD Players')
    title1.grid(row = 0, column = 1, pady = 10)
    club1 = Label(canvas2, text = 'Clubs Name')
    club1.grid(row = 1, column = 0)
    global entry_club
    entry_club = ttk.Combobox(canvas2, width = 12, state = 'readonly')
    entry_club['values'] = ('adama City' , 'bahir dar kenema','debub police','dedebit','defence force','dire dawa','ethiopia bunna','fasil fetema',
             'hawassa','jimma aba jifar','mekelle 70 enderta f.c','shire endeslassie','sidama bunna','st.george','welyata dicha',
             'welewalo adigrat')
    
    entry_club.grid(row = 1, column = 1, pady = 10)
    global entry_player
    player = Label(canvas2, text = 'Player Name')
    player.grid(row = 2, column = 0)
    entry_player = Entry(canvas2)
    entry_player.grid(row = 2, column = 1, pady = 10)
    global entry_jersey
    jersey = Label(canvas2, text = 'Jersey Number')
    jersey.grid(row = 3, column = 0)
    entry_jersey = Entry(canvas2)
    entry_jersey.grid(row = 3, column = 1, pady = 10)
    global entry_age
    age = Label(canvas2, text = 'Age')
    age.grid(row = 4, column = 0)
    entry_age = Entry(canvas2)
    entry_age.grid(row = 4, column = 1, pady = 10)
    global entry_role
    role = Label(canvas2, text = 'Role')
    role.grid(row = 5, column = 0)
    entry_role = ttk.Combobox(canvas2, width = 12, state = 'readonly')
    entry_role['value'] = ('Goal Keeper', 'Middle Defender', 'Right Defender', 'Left Defender', 'Midfielder', 'Right Wing', 'Left Wing', 'Striker')
    entry_role.grid(row = 5, column = 1, padx = 15)
    submit_btn1 = Button(canvas2, text = 'ADD', command = add, bg = 'green')
    submit_btn1.grid(row = 5, column = 2, pady = 10 )
    

    
    
    
 
    canvas1 = Canvas(window, bg = 'white')
    title2 = Label(canvas1, text = 'To Delete Player')
    title2.grid(row = 0, column = 4, pady = 10)
    global entry1_club
    club2 = Label(canvas1, text = 'Clubs Name')
    club2.grid(row = 1, column = 3)
    entry1_club = ttk.Combobox(canvas1, width = 12, state = 'readonly')
    entry1_club['values'] = ('adama City' , 'bahir dar kenema','debub police','dedebit','defence force','dire dawa','ethiopia bunna','fasil fetema',
             'hawassa','jimma aba jifar','mekelle 70 enderta f.c','shire endeslassie','sidama bunna','st.george','welyata dicha',
             'welewalo adigrat')
    
    entry1_club.grid(row = 1, column = 4, pady = 10)
    global entry1_player
    player = Label(canvas1, text = 'Player Name')
    player.grid(row = 2, column = 3)
    entry1_player = Entry(canvas1)
    entry1_player.grid(row = 2, column = 4, padx = 15)
    submit_btn2 = Button(canvas1, text = 'DELETE', command = delete, bg = 'red')
    submit_btn2.grid(row = 2, column = 5, pady = 10 )


    
    canvas3.grid(row = 0,  column = 0)
    canvas2.grid(row = 0, column = 2)
    canvas1.grid(row = 1, columnspan = 3)
    
    
    window.mainloop()
