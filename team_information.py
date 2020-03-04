from tkinter import*
import openpyxl

def AdamaCity():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["adama city"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def  ShireEndeslassie():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["shire endeslassie"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()

def BahirDarKenema():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["bahir dar kenema"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def EthiopiaBunna():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["ethiopia bunna"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
        
def Mekelle():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["mekelle 70 enderta f.c"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()

def Dedebit():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["dedebit"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def DefenceForce():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["defence force"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def DireDawa():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["dire dawa"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()

def FasilKetema():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["fasil ketema"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def Hawassa():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["hawassa"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()

def JimmaAbaJifar():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["jimma aba jifar"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def DebubPolice():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["debub police"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def SidamaBunna():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["sidama bunna"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def George():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["st.george"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def WelyataDicha():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["welyata dicha"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def WolewaloAdigrat():
        window=Tk()
        window.resizable(0,0)
        window.geometry("900x500")
        window.title("Ethiopian Football League  Management System")
        
        sheet_opening=openpyxl.load_workbook("league.xlsx")
        sheet_number=sheet_opening["welewalo adigrat"]
        for i in range(1,40):
            for j in range(1,5):
                cell=sheet_number.cell(row=i,column=j)
                if cell.value !=None:
                    Label(window,text=str(cell.value)).grid(row=i,column=j)
                else:
                    pass
        window.mainloop()
def team():
    global win
    win=Tk()
    win.resizable(0,0)
    win.title("Ethiopian Football League  Management System")
    frame=Frame(win)
    wb = openpyxl.load_workbook('league.xlsx')
    sheet = wb["team"]
    wb.save('league.xlsx')
    Label(frame,text=str(sheet.cell(row=1,column=3).value)).grid(row=1,columnspan=10,pady=10)
    for i in range(2,19):
        for k in range(1,8):
            Label(frame,text=str(sheet.cell(row=i,column=k).value)).grid(row=i,column=k)
    frame.grid(padx=50,pady=50)
    Button(frame,width=30,text="players of Jimma Aba Jifar",command=JimmaAbaJifar).grid(row=3,column=8)
    Button(frame,width=30,text="players of Hawassa",command=Hawassa).grid(row=4,column=8)
    Button(frame,width=30,text="players of Fasil Ketema",command=FasilKetema).grid(row=5,column=8)
    Button(frame,width=30,text="players of Mekelle 70 enderta F.C",command=Mekelle).grid(row=6,column=8)
    Button(frame,width=30,text="players of Dedebit",command=Dedebit).grid(row=7,column=8)
    Button(frame,width=30,text="players of Sidama Bunna",command=SidamaBunna).grid(row=8,column=8)
    Button(frame,width=30,text="players of Wolewalo adigrat",command=WolewaloAdigrat).grid(row=9,column=8)
    Button(frame,width=30,text="players of Adama city",command=AdamaCity).grid(row=10,column=8)
    Button(frame,width=30,text="players of Ethiopia Bunna",command=EthiopiaBunna).grid(row=11,column=8)
    Button(frame,width=30,text="players of Dire Dawa",command=DireDawa).grid(row=12,column=8)
    Button(frame,width=30,text="players of Shire Endeslassie",command=ShireEndeslassie).grid(row=13,column=8)
    Button(frame,width=30,text="players of Welyata Dicha",command=WelyataDicha).grid(row=14,column=8)
    Button(frame,width=30,text="players of Debub Police",command=DebubPolice).grid(row=15,column=8)
    Button(frame,width=30,text="players of Defence Force",command=DefenceForce).grid(row=16,column=8)
    Button(frame,width=30,text="players of st.George",command=George).grid(row=17,column=8)
    Button(frame,width=30,text="players of Bahir Dar Kenema",command=BahirDarKenema).grid(row=18,column=8)
    win.mainloop()